#import xml.etree.ElementTree as ET
from lxml import etree as ET
import sys
import openpyxl
from pathlib import Path
import subprocess
current_file_path = Path(__file__).resolve()
parent_dir = current_file_path.parent.parent
sys.path.append(str(parent_dir))
from name_generator import NameGenerator



def insert_recorder_ids_recursively(element, recorder_id_counter, root):
    for child in element:
        # Skip the activity tag itself
        if child.tag != 'activity':
            # Assign recorderId to the current child element
            child.set('recorderId', f'recorder_id_step_{recorder_id_counter}')
            recorder_id_counter += 1
        # Special handling for 'if' and 'validation' tags
        if child.tag in ['if', 'validation']:
            val_group_ids = child.get('valGroupIds')
            if val_group_ids:
                val_group = root.find(f".//valGroup[@groupId='{val_group_ids}']")
                if val_group is not None:
                    # Find the 'validate' child within 'valGroup' and assign recorderId
                    validate_tag = val_group.find('validate')
                    if validate_tag is not None:
                        validate_tag.set('recorderId', f'recorder_id_step_{recorder_id_counter}')
                        recorder_id_counter += 1

        # Recursively process all child elements
        recorder_id_counter = insert_recorder_ids_recursively(child, recorder_id_counter, root)

    return recorder_id_counter



def insert_recorder_id(data):
    print(f"The activities with the recorder ids to be inserted are: \n{data}")
    recorder_id_counter = 1
    for activity_info in data:
        activity_id = activity_info['activity']
        file_path = activity_info['path']
        # Parse the XML file
        tree = ET.parse(file_path)
        root = tree.getroot()
        # Find the <activity> element with the specified id
        print(f"Searching for activity with ID: {activity_id}")
        activity = root.find(f".//activity[@id='{activity_id}']")
        if activity is not None:
            # Insert recorder IDs recursively starting from the activity element's children
            recorder_id_counter = insert_recorder_ids_recursively(activity, recorder_id_counter, root)
            print(f"Recorder IDs injected for activity '{activity_id}' in '{file_path}'")
            # Save the modified XML back to the file
            tree.write(file_path)
        else:
            print(f"Activity '{activity_id}' not found in '{file_path}'")



def insert_flow(flow_name, flow_desc, flow_path, activities_with_sheets):
    try:
        with open(flow_path, 'r') as file:
            lines = file.readlines()

        for i, line in enumerate(lines):
            if '<flows>' in line:
                lines.insert(i + 1, '\n')
                lines.insert(i + 2, f'\t<flow id="{flow_name}" desc="{flow_desc}" name="{flow_name} : {flow_desc}" summary="{flow_desc}">\n')
                j = 0
                for activity in activities_with_sheets:
                    activity_name = activity['activity']
                    xml_path = activity['path']
                    excelPath = activity['sheets'][0]['excelPath'] if activity['sheets'] else None 
                    sheetName = activity['sheets'][0]['sheetName'] if activity['sheets'] else None
                    if sheetName:
                        lines.insert(i+3+j, f'\t\t<call activity="{activity_name}" xml="{xml_path}" excelPath="{excelPath}" sheetName="{sheetName}"></call>\n')
                    else:
                        lines.insert(i+3+j, f'\t\t<call activity="{activity_name}" xml="{xml_path}"></call>\n')
                    j += 1
                lines.insert(i+3+j, f'\t</flow>\n')
                break  # Exit the loop after inserting
        
        with open(flow_path, 'w') as file:
            file.writelines(lines)
            
    except Exception as e:
        print(f"An error occurred: {e}")


def refactor_for_excel(base_excel_path, activity_path, activity_id):
    nameGenEngine = NameGenerator()
    
    # Load the Excel workbook and select the active sheet
    workbook = openpyxl.load_workbook(base_excel_path)
    sheet = workbook.active

    # Print out the entire content of the Excel sheet for debugging
    print("Excel Sheet Content:")
    for row in sheet.iter_rows(values_only=True):
        print(row)

    # Find the column number for 'recorderId'
    recorder_id_col = None
    for col in sheet.iter_cols(max_row=1, values_only=True):
        if 'recorderId' in col:
            recorder_id_col = col.index('recorderId') + 1  # +1 because Excel is 1-indexed
            break

    if recorder_id_col is None:
        raise ValueError("Column 'recorderId' not found in the Excel sheet.")

    # Create a dictionary to map recorderId to row number and column letters for fast lookup
    recorder_id_map = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))  # Get the header row
    recorder_id_index = header_row.index('recorderId')  # Find the index for the 'recorderId' column
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Starting from row 2, as row 1 is the header
        recorder_id = str(row[recorder_id_index]).strip().lower()  # Use the correct index for recorderId, strip whitespaces, and convert to lowercase
        if recorder_id:  # Ensure recorder_id is not empty
            recorder_id_map[recorder_id] = {
                'row': idx,  # The actual row number in the sheet
                'columns': {header: col_idx + 1 for col_idx, header in enumerate(header_row)}  # Map column names to indices
            }

    print("Recorder IDs found in Excel and their corresponding rows:")
    for recorder_id, info in recorder_id_map.items():
        print(recorder_id, '->', info['row'])

    # Parse the XML file
    tree = ET.parse(activity_path)
    root = tree.getroot()
    
    # Find the activity with the specified activity_id
    activity = root.find(f".//activity[@id='{activity_id}']")
    
    if activity is not None:
        # Iterate over all elements with a 'recorderId' attribute within the activity
        for element in activity.findall(".//*[@recorderId]"):
            recorder_id_xml = element.get('recorderId').strip().lower()  # Strip whitespaces and convert to lowercase
            print(f"Processing element with recorderId: {recorder_id_xml}")  # Debug print

            # Check if this recorderId is present in the Excel file
            if recorder_id_xml in recorder_id_map:
                print(f"Found matching recorderId in Excel: {recorder_id_xml}")  # Debug print
                row_info = recorder_id_map[recorder_id_xml]
                
                # Iterate over the attributes of the XML tag
                for attr_name, attr_value in element.attrib.items():
                    if attr_name != 'recorderId' and attr_name in row_info['columns']:
                        print(f"Refactoring attribute: {attr_name}")  # Debug print
                        
                        # Generate a unique name for the excel variable
                        unique_name = nameGenEngine.get_excel_variable_name()
                        
                        # Create an 'excel' tag with the row, col, and variable attributes
                        excel_tag = ET.Element('excel')
                        excel_tag.set('row', str(row_info['row']))
                        excel_tag.set('col', str(row_info['columns'][attr_name]))
                        excel_tag.set('variable', unique_name)
                        
                        # Insert the excel tag right above the current tag
                        parent_element = element.getparent()
                        if parent_element is not None:
                            parent_element.insert(parent_element.index(element), excel_tag)
                            print(f"Inserted 'excel' tag for {attr_name} before {element.tag}")  # Debug print
                        else:
                            print(f"Parent element not found for {element.tag}")  # Debug print
                        
                        # Replace the attribute value with the variable placeholder
                        element.set(attr_name, f"${{{unique_name}}}")
            else:
                print(f"No matching recorderId found in Excel for: {recorder_id_xml}")  # Debug print
    
    # Save the modified XML back to the file
    tree.write(activity_path, encoding='utf-8', xml_declaration=True)


def find_start_line_number_of_element(root, activity_id):
        print(f"Searching for activity id : {activity_id}...")
        line_number = 1  # Line numbers typically start at 1 
        for elem in root.iter():
            if elem.tag == 'activity' and elem.get('id') == activity_id:
                return line_number  # Return the line number when the flow_id matches
            line_number += 1     
        return None  # Return None if the flow_id is not found


def open_flow(vs_code_path, activity_id, activity_path):
    tree = ET.parse(activity_path)
    root = tree.getroot()
    start_line_number = find_start_line_number_of_element(root, activity_id)

    if start_line_number:
        vs_code_path = vs_code_path.replace('/', '\\')
        command = [vs_code_path, '-g', f"{activity_path}:{start_line_number}"]
        subprocess.Popen(command, shell=True)
    else:
        print(f'Activity ID {activity_id} not found in the XML.')



