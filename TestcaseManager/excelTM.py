import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os


def create_excel(flow_name, base_path):
    # Check if the base_path exists, if not, create it
    if not os.path.exists(base_path):
        os.makedirs(base_path)
        print(f"Created directory: {base_path}")
    # Construct the full file path
    file_name = f"{flow_name}.xlsx"
    file_path = os.path.join(base_path, file_name)
    # Check if the file path is correctly formed
    if not os.path.isabs(file_path):
        print("Error: The file path is not an absolute path.")
        return
    # Define the columns
    columns = ['recorderId', 'tag', 'xpath', 'keyName', 'value', 'condition', 'exists', 'variable1', 'variable2']
    # Create an empty DataFrame with these columns
    df = pd.DataFrame(columns=columns)
    # Save the DataFrame as an Excel file
    try:
        df.to_excel(file_path, index=False)
        print(f"Excel file created at: {file_path}")
    except Exception as e:
        print(f"Error occurred: {e}")

    return file_path



def is_legitimate_path(path):
    # Check if the path is a non-empty string
    if not isinstance(path, str) or not path.strip():
        return False
    # Check if the path exists
    return os.path.exists(path)



import xml.etree.ElementTree as ET
import pandas as pd

def extract_data_and_write_to_excel(data, base_excel_path):
    print(f"Writing the following activities {data} to {base_excel_path}")
    # List to store the extracted data
    extracted_data = []

    for activity_info in data:
        activity_id = activity_info['activity']
        file_path = activity_info['path']

        # Parse the XML file
        tree = ET.parse(file_path)
        root = tree.getroot()

        # Find the <activity> element with the specified id
        activity = root.find(f".//activity[@id='{activity_id}']")
        
        if activity is not None:
            for child in activity:
                # Process 'input', 'if', and 'validation' tags
                if child.tag in ['input', 'if', 'validation']:
                    row_data = {
                        'tag': child.tag,
                        'xpath': child.get('xpath', ''),
                        'value': child.get('value', ''),
                        'recorderId': child.get('recorderId', ''),
                        'variable1': '',
                        'variable2': ''
                    }
                    extracted_data.append(row_data)

                    # Special handling for 'if' and 'validation' tags
                    if child.tag in ['if', 'validation']:
                        val_group_ids = child.get('valGroupIds')
                        if val_group_ids:
                            val_group = root.find(f".//valGroup[@groupId='{val_group_ids}']")
                            if val_group is not None:
                                validate_tag = val_group.find('validate')
                                if validate_tag is not None:
                                    validate_data = {
                                        'tag': 'validate',
                                        'xpath': validate_tag.get('xpath', ''),
                                        'value': validate_tag.get('value', ''),
                                        'recorderId': validate_tag.get('recorderId', ''),
                                        'variable1': validate_tag.get('variable', ''),
                                        'variable2': validate_tag.get('value', ''),
                                        'exists': validate_tag.get('exists', ''),
                                        'condition': validate_tag.get('condition', ''),
                                    }
                                    extracted_data.append(validate_data)

    # Convert the list of dictionaries to a DataFrame and write to Excel
    df = pd.DataFrame(extracted_data)
    df.to_excel(base_excel_path, index=False)
    print(f"Data written to Excel at {base_excel_path}")




def create_duplicates(num_duplicates, base_excel_path):
    sheets = []
    print(f"Creating {num_duplicates} duplicates of sheet1 in {base_excel_path}")
    # Load the workbook and the first sheet
    workbook = load_workbook(base_excel_path)
    sheet_names = workbook.sheetnames
    if not sheet_names:
        print("No sheets found in the Excel file.")
        return
    first_sheet_name = sheet_names[0]
    first_sheet = workbook[first_sheet_name]
    # Duplicate the first sheet num_duplicates times
    for i in range(num_duplicates):
        # Create a new sheet name for each duplicate
        new_sheet_name = f"{first_sheet_name}_Copy{i+1}"
        # Copy the sheet
        new_sheet = workbook.copy_worksheet(first_sheet)
        new_sheet.title = new_sheet_name
        sheets.append(new_sheet_name)
    # Save the workbook
    workbook.save(base_excel_path)
    print(f"Successfully created {num_duplicates} duplicates of '{first_sheet_name}' in '{base_excel_path}'")
    return sheets

